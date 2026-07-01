"""Browse, download and parse the mapping workbooks the agent writes to `output/`.

The pipeline (`run_agent` and the web app's `pipeline_runner`) writes multi-sheet
Excel workbooks to the output directory. This module lets the UI list those files,
serve them for download, and parse them into plain JSON tables for an in-browser
viewer — without re-running any mapping.

Read-only and path-safe: only `*.xlsx` files directly inside the output directory
are ever exposed (no traversal, no subdirectories).
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.utils import project_root

# Header tokens we know the writer emits (see src/output_writer PARENT/ATOMIC_HEADERS).
# Used to locate the real header row inside a sheet that starts with title/subtitle rows.
_KNOWN_HEADERS = {
    "sourceregulation", "sourcecontrolid", "sourcerequirement", "enisacategory",
    "targetregulation", "targetcontrolids", "targetrequirements", "coveragerelationship",
    "coveragelevel", "gap", "detailedgap", "reviewpriority", "sourceatomicid",
    "sourceparentid", "sourceatomicrequirement", "targetatomicids",
    "targetatomicrequirements", "gapclassification", "matchtype", "dimensionscores",
    "gapdimensions", "recommendation", "scoringrationale", "candidatediagnostics",
}

_MAX_ROWS_PER_SHEET = 2000


def output_dir() -> Path:
    """Where the agent writes workbooks. Mirrors src.config default (OUTPUT_DIR or ./output)."""
    value = os.getenv("OUTPUT_DIR", "output")
    p = Path(value)
    return p if p.is_absolute() else project_root() / p


def _norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_listable(path: Path) -> bool:
    name = path.name
    return path.is_file() and name.lower().endswith(".xlsx") and not name.startswith("~$")


def safe_output_path(name: str) -> Path:
    """Resolve `name` to a file inside the output dir, or raise. Blocks path traversal."""
    if not name or "/" in name or "\\" in name or name.startswith("."):
        raise FileNotFoundError("Nom de fichier invalide.")
    base = output_dir().resolve()
    path = (base / name).resolve()
    if base not in path.parents:
        raise FileNotFoundError("Chemin hors du dossier de sortie.")
    if not _is_listable(path):
        raise FileNotFoundError("Fichier introuvable.")
    return path


def _sheet_names(path: Path) -> list[str]:
    try:
        wb = load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def list_outputs() -> list[dict]:
    """List mapping workbooks, most recent first."""
    base = output_dir()
    if not base.exists():
        return []
    files = [p for p in base.iterdir() if _is_listable(p)]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    items: list[dict] = []
    for f in files:
        st = f.stat()
        items.append({
            "name": f.name,
            "size_kb": round(st.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
            "sheets": _sheet_names(f),
        })
    return items


def _detect_header_idx(rows: list[list]) -> int | None:
    """Find the header row: the first early row matching >=3 known headers, else the
    first row with >=2 non-empty cells."""
    for i, row in enumerate(rows[:15]):
        known = sum(1 for c in row if _norm(c) in _KNOWN_HEADERS)
        if known >= 3:
            return i
    for i, row in enumerate(rows):
        if sum(1 for c in row if _clean(c)) >= 2:
            return i
    return 0 if rows else None


def _parse_sheet(ws) -> dict:
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    # Drop fully-empty trailing rows.
    while raw and not any(_clean(c) for c in raw[-1]):
        raw.pop()
    if not raw:
        return {"name": ws.title, "title": "", "headers": [], "rows": [], "total_rows": 0, "truncated": False}

    hdr = _detect_header_idx(raw)
    # Anything above the header row that has content = the sheet title/subtitle.
    title = ""
    for r in raw[:hdr or 0]:
        cell = next((_clean(c) for c in r if _clean(c)), "")
        if cell:
            title = cell
            break

    headers = [_clean(c) for c in raw[hdr]] if hdr is not None else []
    body = raw[(hdr + 1):] if hdr is not None else raw
    total = len(body)
    truncated = total > _MAX_ROWS_PER_SHEET
    body = body[:_MAX_ROWS_PER_SHEET]

    width = max([len(headers)] + [len(r) for r in body]) if body else len(headers)
    headers = (headers + [""] * width)[:width]
    cleaned = [[_clean(c) for c in (r + [None] * width)[:width]] for r in body]

    return {
        "name": ws.title,
        "title": title,
        "headers": headers,
        "rows": cleaned,
        "total_rows": total,
        "truncated": truncated,
    }


def parse_workbook(name: str) -> dict:
    """Parse every sheet of a workbook into JSON tables for the viewer."""
    path = safe_output_path(name)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [_parse_sheet(ws) for ws in wb.worksheets]
    finally:
        wb.close()
    return {"name": name, "sheets": sheets}
