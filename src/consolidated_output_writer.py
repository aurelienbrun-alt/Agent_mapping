"""Excel writer for the consolidated framework output."""
from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .consolidated_framework import ConsolidatedRequirement
from .utils import safe_filename


_NAVY = "17365D"
_LIGHT_GREEN = "E2F0D9"
_LIGHT_BLUE = "D9EAF7"
_LIGHT_YELLOW = "FFF2CC"
_GREY = "F2F2F2"
_WHITE = "FFFFFF"

CONSOLIDATED_HEADERS = [
    "Consolidated ID",
    "ENISA Category",
    "Consolidated Requirement",
    "Origin",
    "Coverage A→B (%)",
    "Source A ID",
    "Source A Requirement",
    "Source B ID(s)",
    "Source B Requirement",
    "B Contribution",
    "Gap A→B",
]

TRACEABILITY_HEADERS = [
    "Consolidated ID",
    "ENISA Category",
    "Origin",
    "Coverage A→B (%)",
    "Source A ID",
    "Source B ID(s)",
]

COL_WIDTHS = [15, 30, 70, 16, 16, 18, 60, 20, 60, 45, 45]


def _header_cell(cell: object, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _origin_fill(origin: str) -> PatternFill:
    color = {
        "A_enriched": _LIGHT_GREEN,
        "B_only": _LIGHT_BLUE,
        "B_supplemental": _LIGHT_YELLOW,
    }.get(origin, _GREY)
    return PatternFill("solid", fgColor=color)


def _data_cell(cell: object, value: object, fill: PatternFill) -> None:
    cell.value = value
    cell.fill = fill
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_consolidated_sheet(ws, requirements: list[ConsolidatedRequirement]) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(CONSOLIDATED_HEADERS, 1):
        _header_cell(ws.cell(row=1, column=col), header)

    for row_idx, req in enumerate(requirements, 2):
        fill = _origin_fill(req.origin)
        values = [
            req.consolidated_id,
            req.category,
            req.consolidated_text,
            req.origin,
            req.coverage_a_to_b,
            req.source_id_a,
            req.source_text_a,
            ", ".join(req.source_ids_b),
            req.source_text_b,
            req.b_contribution,
            req.gap_a_to_b,
        ]
        for col, val in enumerate(values, 1):
            _data_cell(ws.cell(row=row_idx, column=col), val, fill)

    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_traceability_sheet(ws, requirements: list[ConsolidatedRequirement]) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for col, header in enumerate(TRACEABILITY_HEADERS, 1):
        _header_cell(ws.cell(row=1, column=col), header)

    for row_idx, req in enumerate(requirements, 2):
        fill = _origin_fill(req.origin)
        values = [
            req.consolidated_id,
            req.category,
            req.origin,
            req.coverage_a_to_b,
            req.source_id_a,
            ", ".join(req.source_ids_b),
        ]
        for col, val in enumerate(values, 1):
            _data_cell(ws.cell(row=row_idx, column=col), val, fill)

    for col, width in enumerate([15, 30, 16, 16, 20, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_summary_sheet(
    ws,
    requirements: list[ConsolidatedRequirement],
    app_cfg: AppConfig,
) -> None:
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20

    def _row(r: int, label: str, value: object) -> None:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)

    ws.cell(row=1, column=1, value="Consolidated Framework — Summary").font = Font(
        bold=True, size=14, color=_NAVY
    )

    origins = Counter(r.origin for r in requirements)
    _row(3, "Source Framework A", app_cfg.framework_a.name)
    _row(4, "Source Framework B", app_cfg.framework_b.name)
    _row(5, "Total consolidated requirements", len(requirements))
    _row(6, "  A enriched by B (A_enriched)", origins.get("A_enriched", 0))
    _row(7, "  B-only additions (B_only)", origins.get("B_only", 0))
    _row(8, "  B supplemental (B_supplemental)", origins.get("B_supplemental", 0))

    # Coverage distribution for A_enriched
    enriched = [r for r in requirements if r.origin == "A_enriched"]
    if enriched:
        avg_cov = round(sum(r.coverage_a_to_b for r in enriched) / len(enriched))
        _row(10, "Avg coverage A→B (A_enriched requirements)", f"{avg_cov}%")

    # Color legend
    ws.cell(row=12, column=1, value="Legend").font = Font(bold=True)
    for row_idx, (origin, label, color) in enumerate([
        ("A_enriched", "A enriched by B — complete merged requirement", _LIGHT_GREEN),
        ("B_only", "B-only — obligation absent from A (added as-is)", _LIGHT_BLUE),
        ("B_supplemental", "B supplemental — partial B obligation, LLM-verified new", _LIGHT_YELLOW),
    ], start=13):
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=False)


def write_consolidated_workbook(
    app_cfg: AppConfig,
    requirements: list[ConsolidatedRequirement],
    run_id: str,
) -> Path:
    wb = Workbook()

    ws_cons = wb.active
    ws_cons.title = "Consolidated Framework"
    _write_consolidated_sheet(ws_cons, requirements)

    ws_trace = wb.create_sheet("Traceability")
    _write_traceability_sheet(ws_trace, requirements)

    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, requirements, app_cfg)

    now = datetime.now()
    fname = (
        f"consolidated_{safe_filename(app_cfg.framework_a.name)}_"
        f"{safe_filename(app_cfg.framework_b.name)}_"
        f"{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"
    )
    app_cfg.output_dir.mkdir(parents=True, exist_ok=True)
    path = app_cfg.output_dir / fname
    wb.save(path)
    return path
