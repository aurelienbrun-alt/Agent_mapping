from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import AppConfig
from .models import MappingDecision
from .utils import safe_filename


PARENT_HEADERS = [
    "Source regulation",
    "Source control ID",
    "Source requirement",
    "ENISA category",
    "Target regulation",
    "Target control ID(s)",
    "Target requirement(s)",
    "Coverage relationship",
    "Coverage level",
    "Gap",
    "Detailed gap",
    "Review priority",
]

ATOMIC_HEADERS = [
    "Source regulation",
    "Source atomic ID",
    "Source parent ID",
    "Source atomic requirement",
    "ENISA category",
    "Target regulation",
    "Target atomic ID(s)",
    "Target atomic requirement(s)",
    "Coverage relationship",
    "Gap classification",
    "Coverage level",
    "Match type",
    "Dimension scores",
    "Gap dimensions",
    "Detailed gap",
    "Recommendation",
    "Scoring rationale",
    "Candidate diagnostics",
]

BASE_SHEETS = [
    "README",
    "Dashboard",
    "Coverage by Category",
    "Category Quality",
]


def _display_framework_name(name: str) -> str:
    """Return a clean display name for worksheet titles and page headers."""
    text = str(name or "Framework").replace("_", " ").strip()
    text = re.sub(r"\bNIS\s*2\b", "NIS 2", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_sheet_name(name: str) -> str:
    """Excel worksheet names cannot contain brackets, colon, star, question mark, slash or backslash and are limited to 31 chars."""
    text = re.sub(r"[\[\]:*?/\\]", "-", str(name or "Sheet"))
    text = re.sub(r"\s+", " ", text).strip(" '")
    return text or "Sheet"


def _short_framework_name(name: str, max_len: int = 14) -> str:
    """Shorten long regulation names while preserving readable sheet names."""
    text = _clean_sheet_name(_display_framework_name(name))
    if len(text) <= max_len:
        return text
    # Preserve common regulatory acronym + country word when possible.
    words = text.split()
    if len(words) >= 3 and words[0].lower() == "nis" and words[1] == "2":
        candidate = f"NIS2 {words[-1][:2].upper()}"
        if len(candidate) <= max_len:
            return candidate
    if len(words) >= 2:
        candidate = " ".join(words[:2])
        if len(candidate) <= max_len and candidate.upper() not in {"NIS 2"}:
            return candidate
    return text[:max_len].rstrip()


def _make_direction_sheet_name(source_name: str, target_name: str, atomic: bool = False) -> str:
    prefix = "Atomic " if atomic else ""

    if not atomic:
        # For parent sheets, first try the full readable regulation names.
        full_name = f"{_clean_sheet_name(_display_framework_name(source_name))} -> {_clean_sheet_name(_display_framework_name(target_name))}"
        if len(full_name) <= 31:
            return full_name

    left = _short_framework_name(source_name, 10 if atomic else 14)
    right = _short_framework_name(target_name, 10 if atomic else 14)
    name = f"{prefix}{left} -> {right}"
    if len(name) <= 31:
        return _clean_sheet_name(name)

    # If still too long, shorten both sides harder.
    max_side = 7 if atomic else 10
    left = _short_framework_name(source_name, max_side)
    right = _short_framework_name(target_name, max_side)
    name = f"{prefix}{left} -> {right}"
    if len(name) <= 31:
        return _clean_sheet_name(name)
    return _clean_sheet_name(name[:31].rstrip())


def _sheet_names(app_cfg: AppConfig) -> dict[str, str]:
    return {
        "readme": "README",
        "dashboard": "Dashboard",
        "a_parent": _make_direction_sheet_name(app_cfg.framework_a.name, app_cfg.framework_b.name, atomic=False),
        "b_parent": _make_direction_sheet_name(app_cfg.framework_b.name, app_cfg.framework_a.name, atomic=False),
        "coverage": "Coverage by Category",
        "category_quality": "Category Quality",
        "a_atomic": _make_direction_sheet_name(app_cfg.framework_a.name, app_cfg.framework_b.name, atomic=True),
        "b_atomic": _make_direction_sheet_name(app_cfg.framework_b.name, app_cfg.framework_a.name, atomic=True),
    }


def _all_sheet_names(app_cfg: AppConfig, bidirectional: bool = True) -> list[str]:
    names = _sheet_names(app_cfg)
    sheets = [
        names["readme"],
        names["dashboard"],
        names["a_parent"],
    ]
    if bidirectional:
        sheets.append(names["b_parent"])
    sheets.append(names["coverage"])
    sheets.append(names["category_quality"])
    if app_cfg.include_atomic_detail_sheets:
        sheets.append(names["a_atomic"])
        if bidirectional:
            sheets.append(names["b_atomic"])
    return sheets

COLORS = {
    "navy": "17365D",
    "blue": "1F4E78",
    "light_blue": "D9EAF7",
    "green": "70AD47",
    "amber": "FFC000",
    "red": "C00000",
    "light_red": "FCE4D6",
    "light_green": "E2F0D9",
    "light_yellow": "FFF2CC",
    "grey": "F2F2F2",
    "white": "FFFFFF",
    "text": "1F1F1F",
}


# ---------------------------------------------------------------------------
# Final-output coverage bucketing.
# The displayed coverage level is snapped to 0/25/50/75/100 (nearest, half up)
# and the coverage type label is derived deterministically from that bucket.
# This applies ONLY to the final parent sheets and the dashboard; the atomic
# evidence sheets keep the raw granular scores.
# ---------------------------------------------------------------------------
COVERAGE_BUCKETS = (0, 25, 50, 75, 100)

COVERAGE_TYPE_LABELS = {
    0: "Not covered",
    25: "Indirectly covered",
    50: "Partially covered",
    75: "Largely covered",
    100: "Fully covered",
}

_BUCKET_FILL = {
    0:   "FFC7CE",  # red
    25:  "FCE4D6",  # light orange
    50:  "FFF2CC",  # yellow
    75:  "C6EFCE",  # light green
    100: "70D987",  # strong green
}


def _bucket_coverage(value) -> int:
    """Snap a raw coverage value to the nearest of 0/25/50/75/100 (ties round up)."""
    try:
        cov = float(value or 0)
    except Exception:
        cov = 0.0
    cov = max(0.0, min(100.0, cov))
    # key: nearest bucket; on a tie prefer the higher bucket (round half up).
    return int(min(COVERAGE_BUCKETS, key=lambda b: (abs(b - cov), -b)))


def _coverage_type_label(value) -> str:
    return COVERAGE_TYPE_LABELS[_bucket_coverage(value)]


def _bucket_fill(value) -> PatternFill | None:
    color = _BUCKET_FILL.get(_bucket_coverage(value))
    return PatternFill("solid", fgColor=color) if color else None


def _entity_criticality_note(false_targets: list[tuple[str, str]], coverage: int, output_language: str) -> str:
    """Build the 'important entities not covered' gap bullet.

    false_targets is a list of (target_id, target_requirement_text) for the
    selected target requirements flagged Important=False. The note cites what
    those target provisions bring and why it does not apply to important entities.
    """
    if not false_targets:
        return ""
    ids = ", ".join(tid for tid, _ in false_targets if tid) or "—"
    snippet = ""
    for _, text in false_targets:
        text = re.sub(r"\s+", " ", str(text or "")).strip()
        if text:
            snippet = text[:160].rstrip() + ("…" if len(text) > 160 else "")
            break
    if str(output_language or "").lower().startswith("fr"):
        note = (
            f"• Entités importantes : la couverture repose sur l'exigence cible {ids} "
            f"applicable aux seules entités essentielles (Important=False). "
            f"Au recouvrement de {coverage}%, cette obligation n'est pas couverte pour les entités importantes."
        )
        if snippet:
            note += f" La cible indique : « {snippet} »."
    else:
        note = (
            f"• Important entities: coverage relies on target requirement {ids} "
            f"which applies to essential entities only (Important=False). "
            f"At {coverage}% overlap, this obligation is not enforced for important entities."
        )
        if snippet:
            note += f" The target states: “{snippet}”."
    return note


def build_output_filename(app_cfg: AppConfig, now: datetime | None = None) -> Path:
    now = now or datetime.now()
    name = app_cfg.output_filename_pattern.format(
        framework_a=safe_filename(app_cfg.framework_a.name),
        framework_b=safe_filename(app_cfg.framework_b.name),
        date=now.strftime("%Y-%m-%d"),
        time=now.strftime("%H-%M"),
    )
    app_cfg.output_dir.mkdir(parents=True, exist_ok=True)
    candidate = app_cfg.output_dir / name
    if not candidate.exists():
        return candidate
    return app_cfg.output_dir / name.replace(".xlsx", f"_{now.strftime('%S')}.xlsx")


def write_mapping_workbook(
    app_cfg: AppConfig,
    a_to_b: list[MappingDecision],
    b_to_a: list[MappingDecision] | None,
    run_id: str,
    criticality: dict[str, dict[str, bool]] | None = None,
    llm: Any = None,
    logger: Any = None,
) -> Path:
    bidirectional = b_to_a is not None
    if app_cfg.use_output_template and app_cfg.output_template_path.exists():
        wb = load_workbook(app_cfg.output_template_path)
        _reset_sheets(wb, app_cfg, bidirectional=bidirectional)
    else:
        wb = Workbook()
        _reset_sheets(wb, app_cfg, bidirectional=bidirectional)

    sheet_names = _sheet_names(app_cfg)
    source_a_name = _display_framework_name(app_cfg.framework_a.name)
    source_b_name = _display_framework_name(app_cfg.framework_b.name)

    a_parent = _build_parent_rows(a_to_b, source_a_name, source_b_name, app_cfg, criticality)
    b_parent = _build_parent_rows(b_to_a or [], source_b_name, source_a_name, app_cfg, criticality)

    # Action plan synthesis runs on the assembled parent rows so it sees the final
    # Gap text (including the entity-criticality note when enabled).
    if getattr(app_cfg, "enable_action_plan", False) and llm is not None and not app_cfg.dry_run_without_llm:
        from .action_plan import run_action_plan_synthesis
        run_action_plan_synthesis(a_parent + b_parent, app_cfg, llm, logger)

    _write_readme(wb, app_cfg, run_id)
    _write_dashboard(wb, app_cfg, a_parent, b_parent, a_to_b, b_to_a or [], run_id)
    _write_parent_sheet(wb, sheet_names["a_parent"], a_parent, source_a_name, source_b_name, app_cfg)
    if bidirectional:
        _write_parent_sheet(wb, sheet_names["b_parent"], b_parent, source_b_name, source_a_name, app_cfg)
    _write_coverage_by_category(wb, a_parent + b_parent)
    _write_category_quality(wb, a_to_b + (b_to_a or []))

    if app_cfg.include_atomic_detail_sheets:
        _write_atomic_sheet(wb, sheet_names["a_atomic"], a_to_b, source_a_name, source_b_name)
        if bidirectional:
            _write_atomic_sheet(wb, sheet_names["b_atomic"], b_to_a or [], source_b_name, source_a_name)

    _style_tabs(wb, app_cfg)
    output_path = build_output_filename(app_cfg)
    wb.save(output_path)
    return output_path


def _reset_sheets(wb, app_cfg: AppConfig, bidirectional: bool = True) -> None:
    for ws in list(wb.worksheets):
        wb.remove(ws)
    for name in _all_sheet_names(app_cfg, bidirectional=bidirectional):
        wb.create_sheet(name)


def _style_tabs(wb, app_cfg: AppConfig) -> None:
    sheet_names = _sheet_names(app_cfg)
    tab_colors = {
        sheet_names["readme"]: COLORS["grey"],
        sheet_names["dashboard"]: COLORS["navy"],
        sheet_names["a_parent"]: COLORS["blue"],
        sheet_names["b_parent"]: COLORS["blue"],
        sheet_names["coverage"]: COLORS["green"],
        sheet_names["category_quality"]: COLORS["amber"],
        sheet_names["a_atomic"]: COLORS["light_blue"],
        sheet_names["b_atomic"]: COLORS["light_blue"],
    }
    for ws in wb.worksheets:
        color = tab_colors.get(ws.title, COLORS["blue"])
        ws.sheet_properties.tabColor = color


def _write_title(ws, title: str, subtitle: str = "", end_col: int = 10) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True, size=16, color=COLORS["white"])
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws.cell(2, 1).value = subtitle
        ws.cell(2, 1).font = Font(italic=True, size=10, color=COLORS["text"])
        ws.cell(2, 1).fill = PatternFill("solid", fgColor=COLORS["light_blue"])
        ws.row_dimensions[2].height = 22


def _write_readme(wb, app_cfg: AppConfig, run_id: str) -> None:
    ws = wb["README"]
    _write_title(ws, "README - Regulatory mapping output", f"{_display_framework_name(app_cfg.framework_a.name)} ↔ {_display_framework_name(app_cfg.framework_b.name)} | Run {run_id}", 5)
    rows = [
        ["Section", "Description"],
        ["Mapped regulations", f"Framework 1: {_display_framework_name(app_cfg.framework_a.name)}\nFramework 2: {_display_framework_name(app_cfg.framework_b.name)}"],
        ["Purpose", "This workbook summarizes how source requirements are covered by the target regulation. Parent sheets are designed for business/audit reading; atomic sheets preserve the detailed evidence trail."],
        ["Category taxonomy", "Requirements are harmonized to the ENISA category taxonomy before atomization when no framework cache exists. No subcategory scope is used in this edition."],
        ["Coverage level", "Coverage measures how much of the source requirement is covered by the target regulation. It is not a model confidence score."],
        ["Scoring scale", "Coverage uses multiples of 10: 100=exact, 80-90=mostly covered, 40-70=partial (same domain), 10-30=indirect support (different domain), 0=true gap or conflict. Parent scores are aggregated from atomic decisions and may be non-round values such as 43 or 67."],
        ["Relation taxonomy", "The output distinguishes none, true_gap, partial_gap, implementation_detail_gap, indirect_support_gap and conflict_gap. A true_gap is used only when no meaningful target coverage exists."],
        ["Object/action gate", "The gate is a score cap, not a hard rejection. Weak action/object alignment can cap coverage, but the target candidate is kept for traceability when it is relevant."],
        ["Gap column", "The parent Gap column is synthesized by a parent-level LLM judge when enabled. It lists residual gaps by actor, action, object, scope, condition, deadline, evidence, governance and explicitness."],
        ["Keyword matching", "Keyword/BM25 matching and LLM keyword normalization are disabled by default in v3.6. Matching relies on embeddings, structured fields, action/object similarity and category prior."],
        ["Review priority", "High means material missing coverage or low action/object alignment. Medium means partial coverage requiring audit review. Low means mostly covered but not exact."],
        ["Recommended models", "Cost-efficient run: gpt-4.1-nano for atomization/fields/category and text-embedding-3-small for embeddings. Better final judge: gpt-4.1-mini or gpt-5.4-nano if available in your Azure region and budget. Highest assurance: gpt-5.4 or stronger reasoning model for final judge only."],
    ]
    _write_table(ws, 4, 1, rows, table_style=False)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows(min_row=5, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _category_status(avg: float) -> tuple[str, PatternFill, PatternFill]:
    """Return (status label, status fill, avg-coverage fill) for a category row."""
    if avg >= 70:
        status, status_color = "✓ Good", "C6EFCE"
    elif avg >= 50:
        status, status_color = "⚠ Fair", "FFEB9C"
    else:
        status, status_color = "✗ Poor", "FFC7CE"
    cov_color = "C6EFCE" if avg >= 50 else "FFC7CE"
    return status, PatternFill("solid", fgColor=status_color), PatternFill("solid", fgColor=cov_color)


def _short_category(cat: str) -> str:
    text = re.sub(r"^\s*\d+\.\s*", "", str(cat or "")).strip()
    return text[:40].rstrip()


def _dashboard_insights(all_parent: list[dict[str, Any]], by_cat: dict[str, list[dict[str, Any]]], parent_total: int) -> list[str]:
    bodies: list[str] = []
    cat_avgs = sorted(
        ((cat, _avg([i["coverage_level"] for i in items])) for cat, items in by_cat.items()),
        key=lambda x: x[1],
    )
    if cat_avgs:
        worst = cat_avgs[:2]
        parts = " and ".join(f'"{_short_category(c)}" ({a:.0f}%)' for c, a in worst)
        bodies.append(f"PRIORITY CATEGORIES - Focus on {parts} for gap closure")
    partial = sum(1 for r in all_parent if _bucket_coverage(r["coverage_level"]) in (25, 50))
    bodies.append(f"COVERAGE ANALYSIS - {partial} requirement(s) partially covered; review residual evidence requirements")
    high = sum(1 for r in all_parent if r["review_priority"] == "High")
    bodies.append(f"HIGH PRIORITY - {high}/{parent_total} requirement(s) flagged High priority; plan remediation first")
    entity_gaps = sum(1 for r in all_parent if r.get("entity_gap"))
    if entity_gaps:
        bodies.append(
            f"IMPORTANT ENTITIES - {entity_gaps} requirement(s) covered only by essential-entity provisions; "
            "coverage does not hold for important entities"
        )
    gaps = sum(1 for r in all_parent if _bucket_coverage(r["coverage_level"]) == 0)
    bodies.append(f"TRUE GAPS - {gaps} requirement(s) not covered; assess whether compensating controls exist")
    bodies.append("ACTION ITEMS - Develop an implementation roadmap to reach 75%+ coverage across all categories")
    return [f"{i}. {body}" for i, body in enumerate(bodies, start=1)]


def _write_dashboard(wb, app_cfg: AppConfig, a_parent: list[dict[str, Any]], b_parent: list[dict[str, Any]], a_to_b: list[MappingDecision], b_to_a: list[MappingDecision], run_id: str) -> None:
    ws = wb["Dashboard"]
    all_parent = a_parent + b_parent
    all_atomic = a_to_b + b_to_a
    parent_total = len(all_parent)
    atomic_total = len(all_atomic)
    parent_avg = _avg([r["coverage_level"] for r in all_parent])
    atomic_avg = _avg([d.coverage_level for d in all_atomic])

    navy = "1F4788"

    def _section(row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row, 1, text)
        cell.font = Font(bold=True, size=11, color=navy)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Title band + subtitle ---
    ws.merge_cells("A1:H1")
    title = ws.cell(1, 1, "Executive Dashboard - Regulatory Mapping")
    title.font = Font(bold=True, size=16, color=COLORS["white"])
    title.fill = PatternFill("solid", fgColor=navy)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    subtitle = ws.cell(
        2, 1,
        f"{_display_framework_name(app_cfg.framework_a.name)} ↔ {_display_framework_name(app_cfg.framework_b.name)} "
        f"| Generated {datetime.now():%Y-%m-%d}",
    )
    subtitle.font = Font(size=10, color=COLORS["text"])

    # --- KEY METRICS ---
    _section(4, "KEY METRICS")
    metrics = [
        ("Parent Coverage %", f"{parent_avg:.1f}%"),
        ("Parent Requirements", parent_total),
        ("Atomic Coverage %", f"{atomic_avg:.1f}%"),
        ("Total Atomic Decisions", atomic_total),
    ]
    for (label, value), lc in zip(metrics, (1, 3, 5, 7)):
        lab = ws.cell(5, lc, label)
        lab.font = Font(bold=True, size=10)
        val = ws.cell(5, lc + 1, value)
        val.font = Font(bold=True, size=14, color="C65911")
        val.fill = PatternFill("solid", fgColor="FFF2CC")
        val.alignment = Alignment(horizontal="center", vertical="center")

    # --- COVERAGE RELATIONSHIP (bucketed coverage types) ---
    _section(8, "COVERAGE RELATIONSHIP")
    for col, head in enumerate(["Coverage Type", "Count", "Percentage"], start=1):
        hc = ws.cell(9, col, head)
        hc.font = Font(bold=True, size=10, color=COLORS["white"])
        hc.fill = PatternFill("solid", fgColor="4472C4")
        hc.alignment = Alignment(horizontal="center", vertical="center")
    counts = Counter(_coverage_type_label(r["coverage_level"]) for r in all_parent)
    row = 10
    for bucket in (100, 75, 50, 25, 0):
        label = COVERAGE_TYPE_LABELS[bucket]
        cnt = counts.get(label, 0)
        if cnt == 0:
            continue
        pct = (cnt / parent_total * 100) if parent_total else 0
        fill = PatternFill("solid", fgColor=_BUCKET_FILL[bucket])
        ca = ws.cell(row, 1, label); ca.fill = fill
        cb = ws.cell(row, 2, cnt); cb.fill = fill; cb.alignment = Alignment(horizontal="center")
        cp = ws.cell(row, 3, f"{pct:.1f}%"); cp.fill = fill; cp.alignment = Alignment(horizontal="center")
        row += 1

    # --- COVERAGE BY ENISA CATEGORY ---
    cat_section_row = row + 1
    _section(cat_section_row, "COVERAGE BY ENISA CATEGORY")
    head_row = cat_section_row + 1
    for col, head in enumerate(
        ["ENISA Category", "Requirements", "Avg Coverage %", "Status", "High Priority", "Medium Priority"], start=1
    ):
        hc = ws.cell(head_row, col, head)
        hc.font = Font(bold=True, size=10, color=COLORS["white"])
        hc.fill = PatternFill("solid", fgColor="70AD47")
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    by_cat: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in all_parent:
        by_cat[r["domain"]].append(r)
    data_row = head_row + 1
    for cat in sorted(by_cat):
        items = by_cat[cat]
        avg = _avg([i["coverage_level"] for i in items])
        high = sum(1 for i in items if i["review_priority"] == "High")
        medium = sum(1 for i in items if i["review_priority"] == "Medium")
        status, status_fill, cov_fill = _category_status(avg)
        ws.cell(data_row, 1, cat).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(data_row, 2, len(items)).alignment = Alignment(horizontal="center")
        cov_cell = ws.cell(data_row, 3, round(avg, 1)); cov_cell.fill = cov_fill
        cov_cell.alignment = Alignment(horizontal="center")
        st = ws.cell(data_row, 4, status); st.fill = status_fill
        st.alignment = Alignment(horizontal="center")
        ws.cell(data_row, 5, high).alignment = Alignment(horizontal="center")
        ws.cell(data_row, 6, medium).alignment = Alignment(horizontal="center")
        data_row += 1

    # --- KEY INSIGHTS & RECOMMENDATIONS ---
    insights_row = data_row + 1
    _section(insights_row, "KEY INSIGHTS & RECOMMENDATIONS")
    for offset, text in enumerate(_dashboard_insights(all_parent, by_cat, parent_total), start=1):
        line = insights_row + offset
        ws.merge_cells(start_row=line, start_column=1, end_row=line, end_column=8)
        cell = ws.cell(line, 1, text)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 42
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16


def _parent_headers(app_cfg: AppConfig | None = None) -> list[str]:
    base = list(PARENT_HEADERS)  # ends with "Review priority"
    headers = base[:-1]
    if app_cfg and getattr(app_cfg, "enable_action_plan", False):
        headers.append("Action plan")
    headers.append(base[-1])
    if app_cfg and getattr(app_cfg, "enable_entity_criticality", False):
        headers += ["Essential", "Important"]
    return headers


def _write_parent_sheet(wb, sheet_name: str, rows: list[dict[str, Any]], source_name: str, target_name: str, app_cfg: AppConfig | None = None) -> None:
    ws = wb[sheet_name]
    headers = _parent_headers(app_cfg)
    _write_title(ws, f"{source_name} → {target_name}", "Parent-level regulatory mapping. Detailed evidence is available in the atomic detail sheet.", len(headers))
    _write_table(ws, 4, 1, [headers] + [_parent_row_values(r, app_cfg) for r in rows])
    _format_parent_sheet(ws, len(rows) + 4, app_cfg)


def _write_atomic_sheet(wb, sheet_name: str, decisions: list[MappingDecision], source_name: str, target_name: str) -> None:
    ws = wb[sheet_name]
    _write_title(ws, f"Atomic detail - {source_name} → {target_name}", "Detailed evidence trail used to derive the parent-level mapping.", len(ATOMIC_HEADERS))
    matrix = [ATOMIC_HEADERS]
    for d in decisions:
        matrix.append([
            source_name,
            _clean_display_id(d.source_id),
            _clean_display_id(d.source_parent_id),
            d.source_requirement,
            d.source_category,
            target_name,
            ", ".join(d.target_ids),
            "\n---\n".join(d.target_requirements),
            _atomic_relationship(d),
            getattr(d, "gap_type", "") or _gap_type_from_decision(d),
            d.coverage_level,
            d.match_type,
            json_safe(d.dimension_scores),
            "; ".join(d.gap_dimensions),
            d.gap,
            d.recommendation,
            d.scoring_rationale,
            _candidate_diag(d.candidates),
        ])
    _write_table(ws, 4, 1, matrix)
    _format_atomic_sheet(ws, len(matrix) + 3)


def _write_coverage_by_category(wb, rows: list[dict[str, Any]]) -> None:
    ws = wb["Coverage by Category"]
    _write_title(ws, "Coverage by ENISA category", "Aggregated parent-level coverage and review priority by category.", 8)
    matrix = [["ENISA category", "Requirements", "Average coverage", "Gaps", "High priority", "Medium priority", "Low priority", "Exact/Full"]]
    for category, count, cov, high in _coverage_by_category_rows(rows):
        items = [r for r in rows if r["domain"] == category]
        matrix.append([
            category,
            count,
            cov,
            sum(1 for r in items if r["coverage_level"] == 0),
            sum(1 for r in items if r["review_priority"] == "High"),
            sum(1 for r in items if r["review_priority"] == "Medium"),
            sum(1 for r in items if r["review_priority"] == "Low"),
            sum(1 for r in items if r["coverage_level"] == 100),
        ])
    _write_table(ws, 4, 1, matrix)
    _style_table_body(ws, 4, len(matrix), len(matrix[0]))
    if len(matrix) > 1:
        ws.conditional_formatting.add(f"C5:C{len(matrix)+3}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=50, mid_color="FFEB84", end_type="num", end_value=100, end_color="63BE7B"))
    for col, width in zip(range(1, 9), [50, 15, 18, 12, 14, 16, 12, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_category_quality(wb, decisions: list[MappingDecision]) -> None:
    ws = wb["Category Quality"]
    _write_title(ws, "Category harmonization quality", "Traceability of ENISA category harmonization from original framework categories.", 8)
    parent_seen: dict[str, dict[str, Any]] = {}
    for d in decisions:
        # Category harmonization metadata is not on MappingDecision; infer from candidate payload where possible.
        parent_seen[d.source_parent_id or d.source_id] = {
            "Direction": d.direction,
            "Source control ID": d.source_parent_id or d.source_id,
            "ENISA category": d.source_category,
            "Mapping risk": d.mapping_risk,
            "Coverage": d.coverage_level,
        }
    matrix = [["Direction", "Source control ID", "ENISA category", "Mapping risk", "Coverage"]]
    matrix.extend([[v[k] for k in matrix[0]] for v in parent_seen.values()])
    _write_table(ws, 4, 1, matrix)
    _style_table_body(ws, 4, len(matrix), len(matrix[0]))
    for col, width in zip(range(1, 6), [34, 24, 60, 18, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width


def _select_items_for_target_listing(items: list[MappingDecision]) -> list[MappingDecision]:
    """Return the subset of decisions whose targets appear in the parent row.

    When high-quality decisions exist, suppress low-confidence ones to avoid
    an inflated list of target requirements. All items still contribute to
    the coverage average; only the target listing is filtered.
    """
    GOOD_COVERAGE = 50
    good = [d for d in items if d.coverage_level >= GOOD_COVERAGE and d.target_ids]
    if good:
        return good
    with_target = [d for d in items if d.target_ids]
    return with_target or items


def _build_target_listings(
    items: list[MappingDecision],
    min_free: int = 2,
    coverage_ratio: float = 0.50,
    max_targets: int = 5,
) -> tuple[list[str], list[str]]:
    """Select target IDs and requirements with a progressive coverage threshold.

    The first min_free targets are always included. Each subsequent target is
    accepted only when its coverage exceeds average_of_accepted * coverage_ratio.
    Items must be pre-sorted by coverage descending so we can stop as soon as the
    threshold is not met (no later item can do better).
    """
    result_ids: list[str] = []
    result_reqs: list[str] = []
    coverages: list[int] = []

    for d in items:
        if len(result_ids) >= max_targets:
            break
        new_ids = [t for t in (d.target_parent_ids or [_parent_id(x) for x in d.target_ids]) if t and t not in result_ids]
        new_reqs = [r for r in (d.target_parent_requirements or d.target_requirements) if r and r not in result_reqs]
        if not new_ids and not new_reqs:
            continue
        if len(result_ids) < min_free:
            include = True
        else:
            avg = sum(coverages) / len(coverages)
            include = d.coverage_level > avg * coverage_ratio
        if not include:
            break
        slots = max_targets - len(result_ids)
        result_ids.extend(new_ids[:slots])
        result_reqs.extend(new_reqs[:slots])
        coverages.append(d.coverage_level)

    return result_ids, result_reqs


def _build_parent_rows(
    decisions: list[MappingDecision],
    source_name: str,
    target_name: str,
    app_cfg: AppConfig | None = None,
    criticality: dict[str, dict[str, bool]] | None = None,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[MappingDecision]] = defaultdict(list)
    for d in decisions:
        raw_key = d.source_parent_id or _parent_id(d.source_id)
        grouped[_clean_display_id(raw_key)].append(d)

    criticality_enabled = bool(app_cfg and getattr(app_cfg, "enable_entity_criticality", False))
    output_language = getattr(app_cfg, "output_language", "en") if app_cfg else "en"

    rows = []
    for parent_id, items in sorted(grouped.items(), key=lambda kv: kv[0]):
        total = len(items)
        coverage = int(round(sum(d.coverage_level for d in items) / total)) if total else 0
        items_for_targets = sorted(_select_items_for_target_listing(items), key=lambda d: d.coverage_level, reverse=True)
        target_parent_ids, target_parent_requirements = _build_target_listings(items_for_targets)
        source_req = next((d.source_parent_requirement for d in items if d.source_parent_requirement), items[0].source_requirement if items else "")
        domain = next((d.source_category for d in items if d.source_category), "")
        relationship = _parent_relationship(items, coverage)
        gap_type = _parent_gap_type(items, coverage)
        match_type = "None" if not target_parent_ids else "Composite" if len(target_parent_ids) > 1 or len(items) > 1 else items[0].match_type
        gap = _llm_parent_gap_summary(items) or _parent_gap_summary(items, target_name)
        priority = _priority(coverage, items)
        risk = _risk(coverage, items)

        source_essential = True
        source_important = False
        entity_gap = False
        if criticality_enabled:
            src_flags = (criticality or {}).get(parent_id, {})
            source_essential = bool(src_flags.get("essential", True))
            source_important = bool(src_flags.get("important", False))
            # Gap rule: a source applicable to important entities (Important=True)
            # is covered by >=1 target requirement that is NOT applicable to
            # important entities (Important=False) -> flag the residual gap and
            # escalate the review priority.
            if source_important and coverage > 0:
                false_targets = [
                    (tid, req)
                    for tid, req in zip(target_parent_ids, target_parent_requirements)
                    if not (criticality or {}).get(_clean_display_id(tid), {}).get("important", False)
                ]
                if false_targets:
                    note = _entity_criticality_note(false_targets, _bucket_coverage(coverage), output_language)
                    if note:
                        gap = f"{gap}\n\n{note}" if gap else note
                        priority = "High"
                        entity_gap = True

        rows.append({
            "source_regulation": source_name,
            "target_regulation": target_name,
            "source_parent_id": parent_id,
            "source_parent_requirement": source_req,
            "domain": domain,
            "target_parent_ids": target_parent_ids,
            "target_parent_requirements": target_parent_requirements,
            "coverage_relationship": relationship,
            "gap_type": gap_type,
            "coverage_level": coverage,
            "match_type": match_type,
            "gap": gap,
            "review_priority": priority,
            "mapping_risk": risk,
            "source_essential": source_essential,
            "source_important": source_important,
            "entity_gap": entity_gap,
        })
    return rows



def _llm_parent_gap_summary(items: list[MappingDecision]) -> str:
    """Return the LLM-synthesized parent gap summary when available."""
    for item in items:
        value = getattr(item, "parent_gap_summary", "") or ""
        if str(value).strip():
            return str(value).strip()
    return ""


def _parent_gap_type(items: list[MappingDecision], coverage: float) -> str:
    """Classify the parent residual gap from atomic decisions.

    v4 avoids the previous "partial_gap everywhere" behaviour by using
    coverage ratios and the dimensions of residual gaps.
    """
    for item in items:
        value = getattr(item, "parent_gap_type", "") or ""
        if value:
            return _normalize_parent_gap_type(value)
    if not items:
        return "true_gap"
    if coverage >= 99.5:
        return "none"

    scores = [float(getattr(d, "coverage_level", 0) or 0) for d in items]
    total = max(len(scores), 1)
    direct_ratio = sum(1 for s in scores if s >= 40) / total
    high_ratio = sum(1 for s in scores if s >= 80) / total
    indirect_ratio = sum(1 for s in scores if 0 < s < 40) / total
    gap_ratio = sum(1 for s in scores if s <= 0) / total
    types = [_normalize_parent_gap_type(getattr(d, "gap_type", "") or _gap_type_from_decision(d)) for d in items]
    dimensions = {_normalize_dimension(g.get("dimension", "")) for d in items for g in (getattr(d, "gap_items", []) or []) if isinstance(g, dict)}
    implementation_dims = {"evidence", "deadline", "condition", "governance", "explicitness"}
    material_dims = {"actor", "action", "object", "scope"}

    if "conflict_gap" in types:
        return "conflict_gap"
    if gap_ratio >= 0.70 and direct_ratio == 0:
        return "true_gap"
    if direct_ratio == 0 and indirect_ratio > 0:
        return "indirect_support_gap"
    if coverage >= 70 and high_ratio >= 0.50 and dimensions and dimensions.issubset(implementation_dims):
        return "implementation_detail_gap"
    if coverage >= 60 and direct_ratio >= 0.60 and not (dimensions & material_dims):
        return "implementation_detail_gap"
    if coverage >= 40 and direct_ratio >= 0.25:
        return "partial_gap"
    if indirect_ratio > 0 or coverage > 0:
        return "indirect_support_gap"
    return "true_gap"


def _normalize_parent_gap_type(value: str) -> str:
    text = str(value or "").casefold().strip().replace(" ", "_").replace("-", "_")
    aliases = {
        "none": "none",
        "no_gap": "none",
        "true_gap": "true_gap",
        "gap": "true_gap",
        "not_covered": "true_gap",
        "partial": "partial_gap",
        "partial_gap": "partial_gap",
        "implementation": "implementation_detail_gap",
        "implementation_gap": "implementation_detail_gap",
        "implementation_detail": "implementation_detail_gap",
        "implementation_detail_gap": "implementation_detail_gap",
        "indirect": "indirect_support_gap",
        "supportive": "indirect_support_gap",
        "indirect_support": "indirect_support_gap",
        "indirect_support_gap": "indirect_support_gap",
        "conflict": "conflict_gap",
        "conflict_gap": "conflict_gap",
    }
    return aliases.get(text, "partial_gap" if text else "")

def _gap_type_from_decision(d: MappingDecision) -> str:
    if d.coverage_level >= 100:
        return "none"
    relation = str(d.relation_type or "").casefold()
    if "implementation" in relation:
        return "implementation_detail_gap"
    if d.coverage_level >= 40:
        return "partial_gap"
    if d.coverage_level >= 10:
        return "indirect_support_gap"
    return "true_gap"

def _atomic_relationship(d: MappingDecision) -> str:
    """Controlled, filterable taxonomy for atomic coverage relationship."""
    gap_type = _normalize_parent_gap_type(getattr(d, "gap_type", "") or "")
    cov = float(getattr(d, "coverage_level", 0) or 0)
    rel = str(getattr(d, "relation_type", "") or "").casefold()
    if rel == "conflict" or gap_type == "conflict_gap":
        return "Conflicting"
    if rel == "not_covered":
        return "Not covered"
    if cov >= 99.5:
        return "Exact coverage"
    if cov >= 80:
        if gap_type == "implementation_detail_gap":
            return "Implementation details missing"
        return "Mostly covered"
    if cov >= 60:
        return "Largely covered"
    if cov >= 40:
        if gap_type == "implementation_detail_gap":
            return "Implementation details missing"
        return "Partially covered"
    if cov >= 20:
        return "Indirectly covered"
    if cov >= 10:
        return "Very weakly covered"
    return "Not covered"


def _parent_relationship(items: list[MappingDecision], coverage: float) -> str:
    gap_type = _parent_gap_type(items, coverage)
    if gap_type == "conflict_gap":
        return "Conflicting"
    if not items or coverage <= 0:
        return "Not covered"
    if coverage >= 99.5 and gap_type == "none":
        return "Exact coverage"
    if gap_type == "implementation_detail_gap":
        return "Implementation details missing"
    if coverage >= 80:
        return "Mostly covered"
    if coverage >= 60:
        return "Largely covered"
    if coverage >= 40:
        return "Partially covered"
    if coverage >= 20:
        indirect = [d for d in items if 0 < d.coverage_level <= 30]
        if indirect and all(
            str(getattr(d, "relation_type", "") or "").casefold() == "not_covered"
            for d in indirect
        ):
            return "Not covered"
        return "Indirectly covered"
    if coverage >= 10:
        return "Very weakly covered"
    return "Not covered"


def _parent_gap_summary(items: list[MappingDecision], target_name: str) -> str:
    """Build a parent-level residual gap analysis from atomic decisions."""
    if not items:
        return "No mapping decision available."
    if all(d.coverage_level >= 100 for d in items):
        return f"{target_name} fully covers this requirement."

    avg_coverage = round(sum(d.coverage_level for d in items) / max(len(items), 1), 1)
    relationship = _parent_relationship(items, avg_coverage)
    gap_items = _collect_parent_gap_items(items)

    synthesis = _coverage_synthesis_sentence(relationship, avg_coverage, target_name, gap_items)

    if not gap_items:
        return synthesis

    ordered_dimensions = _order_gap_dimensions(item["dimension"] for item in gap_items)
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for item in gap_items:
        grouped[item["dimension"]].append(item)

    bullets: list[str] = []
    for dimension in ordered_dimensions:
        ordered_items = sorted(grouped[dimension], key=lambda x: _severity_rank(x.get("severity", "moderate")))
        for item in ordered_items:
            text = item.get("gap", "").strip()
            if text:
                bullets.append(f"• {text}")

    if not bullets:
        return synthesis
    return synthesis + "\n\n" + "\n".join(bullets)


def _collect_parent_gap_items(items: list[MappingDecision]) -> list[dict[str, str]]:
    collected: list[dict[str, str]] = []
    for d in sorted(items, key=lambda x: (x.coverage_level, x.source_id)):
        if d.coverage_level >= 100:
            continue
        structured = getattr(d, "gap_items", None) or []
        if isinstance(structured, list):
            for raw in structured:
                item = _normalize_gap_item(raw)
                if item:
                    collected.append(item)
        if not structured and d.gap:
            dimensions = d.gap_dimensions or []
            snippets = _split_gap_points(d.gap)
            for snippet in snippets:
                dim = _infer_gap_dimension(snippet, dimensions)
                sev = _severity_from_coverage(d.coverage_level)
                collected.append({"dimension": dim, "gap": snippet, "severity": sev, "target_coverage": ""})
    return _dedupe_gap_items(collected)


def _normalize_gap_item(raw: Any) -> dict[str, str] | None:
    if not isinstance(raw, dict):
        return None
    gap = str(raw.get("gap") or raw.get("missing_element") or raw.get("residual_gap") or "").strip()
    if not gap or _is_generic_gap_text(gap):
        return None
    dimension = _normalize_dimension(str(raw.get("dimension") or raw.get("missing_dimension") or "explicitness"))
    severity = _normalize_severity(str(raw.get("severity") or "moderate"))
    target_coverage = str(raw.get("target_coverage") or raw.get("coverage_note") or "").strip()
    return {
        "dimension": dimension,
        "gap": _clean_gap_sentence(gap),
        "severity": severity,
        "target_coverage": _clean_gap_sentence(target_coverage) if target_coverage else "",
    }


def _split_gap_points(text: str) -> list[str]:
    cleaned = re.sub(r"\s+", " ", str(text or "")).strip()
    cleaned = re.sub(r"Coverage downgraded by object/action gate:", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"Coverage capped by object/action gate[^.]*\.\s*", "", cleaned, flags=re.IGNORECASE)
    raw_parts = re.split(r"(?:\n+|\s*[•▪]\s*|(?<=[.;])\s+)", cleaned)
    out = []
    for part in raw_parts:
        part = _clean_gap_sentence(part)
        if not part or _is_generic_gap_text(part):
            continue
        out.append(part)
    return out


def _clean_gap_sentence(text: str) -> str:
    text = re.sub(r"\s+", " ", str(text or "")).strip(" -\n\t;")
    text = re.sub(r"^\w+[\w./-]*#\d+\s*\([^)]*\)\s*->\s*[^:]+:\s*", "", text)
    text = re.sub(r"^(gap|residual gap|missing element)\s*:\s*", "", text, flags=re.IGNORECASE)
    text = text.strip(" -\n\t;")
    if text and text[-1] not in ".!?":
        text += "."
    return text


def _is_generic_gap_text(text: str) -> bool:
    t = re.sub(r"\s+", " ", str(text or "")).casefold().strip()
    generic_fragments = [
        "target framework does not fully cover the source requirement",
        "review detailed analysis in the atomic sheet",
        "manual review required",
        "no selected target candidate",
        "no sufficient coverage identified",
        "no sufficient match passed retrieval",
        "obvious gap shortcut",
        "coverage capped by object/action gate",
        "coverage downgraded by object/action gate",
        "candidate is kept for traceability",
        "not forced to gap",
    ]
    return any(fragment in t for fragment in generic_fragments)


def _dedupe_gap_items(items: list[dict[str, str]]) -> list[dict[str, str]]:
    unique: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    for item in items:
        gap = item.get("gap", "")
        if not gap:
            continue
        key = _gap_dedupe_key(item.get("dimension", ""), gap)
        if key in seen_keys:
            continue
        # Drop near-duplicates based on high token overlap within the same dimension.
        tokens = set(tokenize_for_gap(gap))
        duplicate = False
        for existing in unique:
            if existing.get("dimension") != item.get("dimension"):
                continue
            ex_tokens = set(tokenize_for_gap(existing.get("gap", "")))
            if tokens and ex_tokens and len(tokens & ex_tokens) / max(len(tokens | ex_tokens), 1) >= 0.78:
                duplicate = True
                break
        if duplicate:
            continue
        seen_keys.add(key)
        unique.append(item)
    return unique


def tokenize_for_gap(text: str) -> list[str]:
    return [t for t in re.findall(r"[a-zA-ZÀ-ÿ0-9]+", str(text).casefold()) if len(t) > 2]


def _gap_dedupe_key(dimension: str, gap: str) -> str:
    tokens = tokenize_for_gap(gap)
    return f"{_normalize_dimension(dimension)}|{' '.join(tokens[:16])}"


def _normalize_dimension(value: str) -> str:
    text = str(value or "").strip().casefold()
    aliases = {
        "actors": "actor",
        "role": "actor",
        "responsibility": "governance",
        "responsibilities": "governance",
        "process": "governance",
        "procedure": "governance",
        "frequency": "deadline",
        "timing": "deadline",
        "proof": "evidence",
        "documentation": "evidence",
        "explicit": "explicitness",
        "specificity": "explicitness",
        "granularity": "explicitness",
    }
    allowed = {"actor", "action", "object", "scope", "condition", "deadline", "evidence", "governance", "explicitness", "control_type"}
    text = aliases.get(text, text)
    return text if text in allowed else "explicitness"


def _infer_gap_dimension(text: str, preferred: list[str]) -> str:
    for dim in preferred:
        norm = _normalize_dimension(dim)
        if norm:
            return norm
    t = str(text or "").casefold()
    rules = [
        ("actor", ["actor", "role", "responsible", "entity", "personnel", "third party", "prestataire"]),
        ("evidence", ["evidence", "record", "documented", "proof", "trace", "journal", "documenté", "preuve"]),
        ("deadline", ["deadline", "frequency", "periodic", "annual", "monthly", "timeframe", "délai", "périodique", "annuel"]),
        ("scope", ["scope", "perimeter", "critical", "essential", "ot", "ict", "all systems", "périmètre", "critique"]),
        ("condition", ["condition", "where", "when", "if", "risk-based", "as necessary", "si", "lorsque"]),
        ("governance", ["policy", "procedure", "governance", "approval", "responsibility", "process", "politique", "procédure"]),
        ("action", ["define", "implement", "review", "test", "monitor", "notify", "maintain", "mettre en œuvre", "surveiller", "notifier", "tester"]),
        ("object", ["asset", "system", "account", "access", "data", "backup", "logs", "ressource", "compte", "données"]),
    ]
    for dim, keywords in rules:
        if any(k in t for k in keywords):
            return dim
    return "explicitness"


def _normalize_severity(value: str) -> str:
    text = str(value or "").casefold().strip()
    if text in {"critical", "high", "material", "major"}:
        return "material"
    if text in {"low", "minor", "residual"}:
        return "minor"
    return "moderate"


def _severity_rank(value: str) -> int:
    return {"material": 0, "moderate": 1, "minor": 2}.get(_normalize_severity(value), 1)


def _severity_from_coverage(coverage: int | float) -> str:
    try:
        cov = float(coverage)
    except Exception:
        cov = 0
    if cov <= 25:
        return "material"
    if cov <= 50:
        return "moderate"
    return "minor"


def _order_gap_dimensions(dimensions) -> list[str]:
    order = ["action", "object", "scope", "actor", "condition", "deadline", "evidence", "governance", "control_type", "explicitness"]
    dims = {_normalize_dimension(d) for d in dimensions}
    return [d for d in order if d in dims] + sorted(dims - set(order))


def _coverage_synthesis_sentence(relationship: str, coverage: float, target_name: str, gap_items: list[dict[str, str]]) -> str:
    material = sum(1 for item in gap_items if item.get("severity") == "material")
    moderate = sum(1 for item in gap_items if item.get("severity") == "moderate")
    if coverage <= 0 or relationship == "Not covered":
        return f"{target_name} does not identify a direct equivalent obligation for the source requirement."
    if relationship == "Indirectly covered":
        return f"{target_name} provides related or supportive provisions, but they do not establish equivalent direct coverage."
    if relationship == "Partially covered":
        return f"{target_name} partially covers the source requirement, with {material + moderate} substantive residual difference(s)."
    if relationship == "Mostly covered":
        return f"{target_name} mostly covers the source requirement, with residual differences requiring interpretation or complementary evidence."
    return f"{target_name} covers the main objective, but residual differences remain."



def _priority(coverage: float, items: list[MappingDecision]) -> str:
    if coverage <= 25 or any(d.equivalence_level == "Gap" for d in items):
        return "High"
    if coverage < 75 or any((d.mapping_risk or "").lower() == "high" for d in items):
        return "Medium"
    if coverage < 100:
        return "Low"
    return "None"


def _risk(coverage: float, items: list[MappingDecision]) -> str:
    risks = {str(d.mapping_risk or "").lower() for d in items}
    if "high" in risks or coverage <= 25:
        return "High"
    if "medium" in risks or coverage < 75:
        return "Medium"
    if coverage < 100:
        return "Low"
    return "None"


def _split_gap_text(gap: str) -> tuple[str, str]:
    """Split gap text into (synthesis sentence, bullet points).

    The synthesis is the first non-empty paragraph before any bullet lines.
    The detailed section collects all lines starting with '•'.
    """
    lines = str(gap or "").splitlines()
    synthesis_lines: list[str] = []
    detail_lines: list[str] = []
    in_bullets = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("•"):
            in_bullets = True
            detail_lines.append(stripped)
        elif not in_bullets:
            if stripped:
                synthesis_lines.append(stripped)
    return " ".join(synthesis_lines).strip(), "\n".join(detail_lines).strip()


def _parent_row_values(r: dict[str, Any], app_cfg: AppConfig | None = None) -> list[Any]:
    gap_synthesis, gap_detail = _split_gap_text(r["gap"])
    # Final output: snap coverage to 0/25/50/75/100 and derive the coverage type
    # label deterministically from that bucket so the two columns stay consistent.
    coverage_display = _bucket_coverage(r["coverage_level"])
    relationship_display = _coverage_type_label(r["coverage_level"])
    values = [
        r["source_regulation"],
        _clean_display_id(r["source_parent_id"]),
        r["source_parent_requirement"],
        r["domain"],
        r["target_regulation"],
        ", ".join(r["target_parent_ids"]),
        "\n---\n".join(r["target_parent_requirements"]),
        relationship_display,
        coverage_display,
        gap_synthesis,
        gap_detail,
    ]
    if app_cfg and getattr(app_cfg, "enable_action_plan", False):
        values.append(r.get("action_plan", ""))
    values.append(r["review_priority"])
    if app_cfg and getattr(app_cfg, "enable_entity_criticality", False):
        values.append("True" if r.get("source_essential", True) else "False")
        values.append("True" if r.get("source_important", False) else "False")
    return values


def _write_table(ws, start_row: int, start_col: int, matrix: list[list[Any]], table_style: bool = True) -> None:
    if not matrix:
        return
    for r, row in enumerate(matrix, start=start_row):
        for c, value in enumerate(row, start=start_col):
            ws.cell(r, c).value = value
    _style_table_body(ws, start_row, len(matrix), len(matrix[0]))


def _style_table_body(ws, start_row: int, row_count: int, col_count: int) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_row = start_row
    for c in range(1, col_count + 1):
        cell = ws.cell(header_row, c)
        cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + row_count - 1, min_col=1, max_col=col_count):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(col_count)}{start_row + row_count - 1}"


_COVERAGE_BG = {
    0:   "FFC7CE",  # red          — Not covered
    10:  "FFCFC9",  # light red    — Minimal coverage
    20:  "FDDBD1",  # salmon       — Very indirect
    30:  "FDDFC0",  # light orange — Indirectly covered
    40:  "FFF0C0",  # pale yellow  — Weakly covered
    50:  "FFF2CC",  # yellow       — Partially covered
    60:  "F5F5C0",  # yellow-green — Well covered
    70:  "E2F0D0",  # light green  — Largely covered
    80:  "C6EFCE",  # green        — Mostly covered
    90:  "A8E6B8",  # medium green — Almost full
    100: "70D987",  # strong green — Exact coverage
}


def _coverage_fill(value) -> PatternFill | None:
    try:
        cov = int(round(float(value or 0)))
    except Exception:
        return None
    # Snap to nearest 10% zone.
    zone = min(_COVERAGE_BG.keys(), key=lambda x: abs(x - cov))
    color = _COVERAGE_BG.get(zone)
    return PatternFill("solid", fgColor=color) if color else None


_PARENT_COL_WIDTHS = {
    "Source regulation": 24,
    "Source control ID": 22,
    "Source requirement": 70,
    "ENISA category": 48,
    "Target regulation": 24,
    "Target control ID(s)": 28,
    "Target requirement(s)": 70,
    "Coverage relationship": 22,
    "Coverage level": 14,
    "Gap": 55,
    "Detailed gap": 60,
    "Action plan": 55,
    "Review priority": 18,
    "Essential": 12,
    "Important": 12,
}


def _format_parent_sheet(ws, max_row: int, app_cfg: AppConfig | None = None) -> None:
    # Column positions are resolved by header name (row 4) so optional columns
    # (Action plan, Essential, Important) don't break fixed indices.
    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _PARENT_COL_WIDTHS.get(header, 20)

    cov_rel_c = col.get("Coverage relationship")
    cov_lvl_c = col.get("Coverage level")
    priority_c = col.get("Review priority")
    essential_c = col.get("Essential")
    important_c = col.get("Important")
    priority_fill = {
        "High": COLORS["light_red"],
        "Medium": COLORS["light_yellow"],
        "Low": COLORS["light_green"],
    }

    for r in range(5, max_row + 1):
        ws.row_dimensions[r].height = 60
        # Coverage relationship + Coverage level share the 0/25/50/75/100 palette.
        if cov_lvl_c:
            cov_fill = _bucket_fill(ws.cell(r, cov_lvl_c).value)
            if cov_fill:
                if cov_rel_c:
                    ws.cell(r, cov_rel_c).fill = cov_fill
                ws.cell(r, cov_lvl_c).fill = cov_fill
        if priority_c:
            color = priority_fill.get(str(ws.cell(r, priority_c).value or ""))
            if color:
                ws.cell(r, priority_c).fill = PatternFill("solid", fgColor=color)
        for c in (essential_c, important_c):
            if c:
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
        if important_c and str(ws.cell(r, important_c).value or "") == "True":
            ws.cell(r, important_c).fill = PatternFill("solid", fgColor=COLORS["light_blue"])


def _format_atomic_sheet(ws, max_row: int) -> None:
    widths = [24, 24, 22, 70, 48, 24, 28, 70, 22, 22, 16, 18, 34, 30, 70, 58, 50, 70]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    if max_row >= 5:
        ws.conditional_formatting.add(f"K5:K{max_row}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=50, mid_color="FFEB84", end_type="num", end_value=100, end_color="63BE7B"))


def _coverage_by_category_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_category[row["domain"]].append(row)
    result = []
    for cat, items in sorted(by_category.items()):
        result.append([
            cat,
            len(items),
            round(_avg([i["coverage_level"] for i in items]), 1),
            sum(1 for i in items if i["review_priority"] == "High"),
        ])
    return result


def _avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _candidate_diag(candidates: list[dict[str, Any]], max_items: int = 5) -> str:
    parts = []
    for c in (candidates or [])[:max_items]:
        parts.append(
            f"{c.get('rank')}. {c.get('candidate_id')} | combined={c.get('combined_score')} | action_object={c.get('action_object_score')} | semantic={c.get('semantic_score')} | gate={c.get('hard_gate')}"
        )
    return "\n".join(parts)


def _unique(values) -> list[str]:
    out: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in out:
            out.append(text)
    return out


def _dedupe_text(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = re.sub(r"\s+", " ", str(value or "")).strip()
        key = text.casefold()
        if text and key not in seen:
            seen.add(key)
            out.append(text)
    return out


def json_safe(value: Any) -> str:
    if value in (None, "", {}, []):
        return ""
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return str(value)


def _parent_id(atomic_id: str) -> str:
    return str(atomic_id).rsplit("#", 1)[0]


def _clean_display_id(id_str: str) -> str:
    """Strip internal __row_N suffixes added when source data has duplicate IDs."""
    return re.sub(r"__row_\d+$", "", str(id_str or ""))
