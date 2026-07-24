from __future__ import annotations

from typing import TYPE_CHECKING

import pandas as pd

from .config import FrameworkConfig, AppConfig
from .models import RequirementRow
from .utils import normalize_text, normalize_category

if TYPE_CHECKING:
    from .logging_utils import JsonlRunLogger


def _configured_columns(cfg: FrameworkConfig) -> dict[str, str]:
    """Return only columns explicitly configured in .env.

    Empty values such as A_SUBCATEGORY_COLUMN= mean "this field does not exist in
    the Excel file" and must not be validated against df.columns.
    """
    return {
        "id": normalize_text(cfg.id_column),
        "title": normalize_text(cfg.title_column),
        "requirement": normalize_text(cfg.requirement_column),
        "category": normalize_text(cfg.category_column),
        "subcategory": normalize_text(cfg.subcategory_column),
    }


def _cell(row, column_name: str) -> str:
    if not column_name:
        return ""
    return normalize_text(row.get(column_name))


_TRUE_TOKENS = {"true", "vrai", "1", "yes", "oui", "y", "o", "x", "✓", "✔", "essential", "important"}
_FALSE_TOKENS = {"false", "faux", "0", "no", "non", "n", "", "na", "n/a", "-"}


def _cell_bool(row, column_name: str, default: bool) -> bool:
    """Parse a True/False criticality cell. Empty/unknown values fall back to default."""
    if not column_name:
        return default
    raw = normalize_text(row.get(column_name)).strip().lower()
    if raw in _TRUE_TOKENS:
        return True
    if raw in _FALSE_TOKENS:
        return False
    return default


def _warn_if_id_column_numeric(cfg: FrameworkConfig, id_column: str) -> None:
    """Detect a control-ID column Excel stored as NUMBERS rather than text.

    A cell containing "5.10" typed into a General-formatted column is silently
    stored as the float 5.1 and collides with a real "5.1" control — the trailing
    zero is unrecoverable from the value alone (found on the ISO 27001 catalog:
    5.10/5.20/5.30/7.10/8.10/8.20/8.30 all collided with 5.1/5.2/5.3/7.1/8.1/8.2/8.3,
    corrupting 7 of 93 controls). `dtype=str` below only stringifies whatever
    pandas already parsed, so it cannot recover a value already destroyed upstream.
    """
    if not id_column:
        return
    # Inspect the RAW Excel cell types with openpyxl, not pandas' inferred dtype.
    # `pd.read_excel` without dtype=str re-coerces a numeric-looking *text* column
    # (e.g. a correctly-fixed "5.10"/"5.1" column) back to float64 and re-truncates
    # it, which would fire this warning on a file that is actually fine. Only a cell
    # Excel stored as a NUMBER (data_type == 'n') can lose a trailing zero.
    try:
        from openpyxl import load_workbook

        wb = load_workbook(cfg.file, read_only=True, data_only=True)
        ws = wb[cfg.sheet_name] if cfg.sheet_name else wb.worksheets[0]
        rows_iter = ws.iter_rows()
        header = [normalize_text(c.value) for c in next(rows_iter)]
        if id_column not in header:
            wb.close()
            return
        ci = header.index(id_column)
        seen: set[str] = set()
        dupes_set: set[str] = set()
        numeric_present = False
        for row in rows_iter:
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell.value is None:
                continue
            if cell.data_type == "n":  # stored as a number, not text
                numeric_present = True
            key = str(cell.value).strip()
            if key in seen:
                dupes_set.add(key)
            seen.add(key)
        wb.close()
    except Exception:
        return  # the real read below will raise a clearer error if the file is bad
    # Only warn on an ACTUAL collision produced by numeric storage — not merely
    # because the column happens to be numeric-typed (plain "3.1" clause numbers
    # are numeric and never collide), and not for a correctly text-typed column.
    dupes = sorted(dupes_set)
    if dupes and numeric_present:
        print(
            f"[WARNING] {cfg.file.name}: ID column '{id_column}' is stored as NUMBERS in Excel, "
            f"and {len(dupes)} value(s) repeat after stringifying ({', '.join(dupes[:5])}"
            f"{'...' if len(dupes) > 5 else ''}). A repeat here likely means a trailing-zero ID "
            "(e.g. 5.10) collided with a shorter one (5.1) and the difference is unrecoverable. "
            "Fix: format that column as Text in Excel (or prefix values with an apostrophe, "
            "'5.10) before re-running.",
            flush=True,
        )


def read_framework_excel(
    cfg: FrameworkConfig,
    app_cfg: AppConfig,
    logger: "JsonlRunLogger | None" = None,
) -> list[RequirementRow]:
    if not cfg.file.exists():
        raise FileNotFoundError(f"Excel file not found: {cfg.file}")

    columns = _configured_columns(cfg)
    _warn_if_id_column_numeric(cfg, columns["id"])
    df = pd.read_excel(cfg.file, sheet_name=cfg.sheet_name or 0, dtype=str)

    # The requirement column is the only truly mandatory field.
    # ID, title, category and subcategory may be left empty in .env.
    if not columns["requirement"]:
        raise ValueError(
            f"Missing requirement column configuration for {cfg.file.name}. "
            "Set A_REQUIREMENT_COLUMN or B_REQUIREMENT_COLUMN in .env."
        )

    missing = [
        col
        for col in columns.values()
        if col and col not in df.columns
    ]
    if missing:
        available = list(df.columns)
        raise ValueError(
            f"Missing columns in {cfg.file.name}: {missing}. Check .env column names. "
            f"Available columns are: {available}"
        )

    # Entity criticality columns are validated only when the feature is enabled.
    essential_col = normalize_text(cfg.essential_column) if app_cfg.enable_entity_criticality else ""
    important_col = normalize_text(cfg.important_column) if app_cfg.enable_entity_criticality else ""
    if app_cfg.enable_entity_criticality:
        crit_missing = [c for c in (essential_col, important_col) if c and c not in df.columns]
        if crit_missing:
            raise ValueError(
                f"ENABLE_ENTITY_CRITICALITY=true but missing column(s) in {cfg.file.name}: {crit_missing}. "
                f"Set A_ESSENTIAL_COLUMN/A_IMPORTANT_COLUMN (and B_*) in .env or disable the feature. "
                f"Available columns are: {list(df.columns)}"
            )

    if app_cfg.max_requirements_per_framework > 0:
        df = df.head(app_cfg.max_requirements_per_framework)

    rows: list[RequirementRow] = []
    seen_ids: set[str] = set()
    duplicate_ids: list[str] = []
    for idx, row in df.iterrows():
        requirement = _cell(row, columns["requirement"])
        source_id = _cell(row, columns["id"]) or f"row_{idx + 2}"
        if not requirement:
            continue
        if source_id in seen_ids:
            # A control framework must not carry two controls under the same ID.
            # The usual cause is an ID column stored as *numbers*: a control ending
            # in a zero (5.30, 8.20, ...) is held as the float 5.3 / 8.2, collapses
            # onto the real 5.3 / 8.2, and is silently renamed here to
            # "<id>__row_N" — which corrupts every mapping that should have reached
            # the lost control. Record it so we can warn instead of shipping a
            # wrong workbook. Fix the source with tools/fix_iso_catalog_ids.py.
            duplicate_ids.append(source_id)
            source_id = f"{source_id}__row_{idx + 2}"
        seen_ids.add(source_id)

        category = _cell(row, columns["category"])
        subcategory = _cell(row, columns["subcategory"])

        # Essential is always True in the source data; Important defaults to False when absent.
        essential = _cell_bool(row, essential_col, default=True)
        important = _cell_bool(row, important_col, default=False)

        rows.append(
            RequirementRow(
                framework=cfg.name,
                source_id=source_id,
                title=_cell(row, columns["title"]),
                requirement=requirement,
                category=category,
                category_key=normalize_category(
                    category,
                    case_sensitive=app_cfg.category_case_sensitive,
                    trim_spaces=app_cfg.category_trim_spaces,
                ),
                subcategory=subcategory,
                row_number=int(idx) + 2,
                essential=essential,
                important=important,
            )
        )

    if duplicate_ids:
        offenders = sorted(set(duplicate_ids))
        message = (
            f"{cfg.file.name}: {len(offenders)} duplicate source ID(s) after read: {offenders}. "
            "A likely cause is IDs stored as numbers (controls ending in 0 such as 5.30/8.20 "
            "collapse onto 5.3/8.2). Store the ID column as text, e.g. run "
            "tools/fix_iso_catalog_ids.py. The duplicates were kept under '<id>__row_N' but the "
            "output IDs will be ambiguous."
        )
        if logger is not None:
            logger.event(
                "framework.read.duplicate_ids",
                status="warning",
                framework=cfg.name,
                file=str(cfg.file),
                duplicates=offenders,
                hint=message,
            )
        else:  # keep the signal even when called outside the pipeline (tools, tests)
            import warnings

            warnings.warn(message, stacklevel=2)

    return rows
