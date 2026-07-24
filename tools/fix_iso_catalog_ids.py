from __future__ import annotations

"""Repair ISO catalog control IDs that Excel stored as numbers.

Root cause: when the ID column is stored as a *number*, controls whose ID ends
in a zero (5.10, 5.20, 5.30, 7.10, 8.10, 8.20, 8.30) are held as the floats
5.1/5.2/5.3/7.1/8.1/8.2/8.3 and collide with the real low-numbered control. The
pipeline then dedups them to ``X.Y__row_N`` and the workbook shows an ambiguous
ID (e.g. a business-continuity requirement mapped to "5.3 Segregation of duties"
instead of "5.30 ICT readiness for business continuity").

This tool rewrites the ID column as canonical *text* IDs, reconstructed from the
row position within each control family. It is deterministic and idempotent: it
recovers the correct IDs even from a fully-numeric (regressed) file, and leaves
an already-correct file semantically unchanged. It refuses to touch a file whose
per-family control counts do not match the ISO/IEC 27002:2022 Annex A catalog,
so it never silently mis-numbers a different framework.

Usage (from project root):
    python tools/fix_iso_catalog_ids.py --check           # report only, no write
    python tools/fix_iso_catalog_ids.py                   # repair data/ISO.xlsx
    python tools/fix_iso_catalog_ids.py --file data/ISO.xlsx --column ID
"""

import argparse
import shutil
from pathlib import Path
import sys

import openpyxl

ROOT = Path(__file__).resolve().parents[1]

# ISO/IEC 27002:2022 Annex A: number of controls per theme family.
EXPECTED_FAMILY_COUNTS = {"5": 37, "6": 8, "7": 14, "8": 34}


def _family_of(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or "." not in text:
        return None
    return text.split(".", 1)[0].strip()


def repair_workbook(path: Path, column: str, *, check_only: bool) -> int:
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]

    header = [c.value for c in ws[1]]
    if column not in header:
        print(f"ERROR: column {column!r} not found. Header = {header}")
        return 2
    col_idx = header.index(column) + 1  # openpyxl is 1-based

    # Walk data rows in order, reconstruct canonical IDs family-by-family.
    counters: dict[str, int] = {}
    plan: list[tuple[int, str, str]] = []  # (excel_row, old_repr, new_id)
    family_counts: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        family = _family_of(cell.value)
        if family is None:
            continue  # blank / spacer row
        counters[family] = counters.get(family, 0) + 1
        family_counts[family] = counters[family]
        new_id = f"{family}.{counters[family]}"
        old_repr = f"{cell.value!r}({cell.data_type})"
        plan.append((r, old_repr, new_id))

    # Safety gate: only rewrite when the structure matches the known ISO catalog.
    if family_counts != EXPECTED_FAMILY_COUNTS:
        print("ERROR: per-family control counts do not match ISO/IEC 27002:2022.")
        print(f"  found    = {family_counts}")
        print(f"  expected = {EXPECTED_FAMILY_COUNTS}")
        print("  Refusing to renumber. Pass --force to override (advanced).")
        return 3

    changed = [(r, old, new) for (r, old, new) in plan if not _same(ws, r, col_idx, new)]
    lost = [new for (_, _, new) in plan if new.endswith("0")]
    print(f"File: {path}")
    print(f"  data rows with an ID: {len(plan)} | families: {family_counts}")
    print(f"  trailing-zero controls that must be text: {sorted(lost)}")
    print(f"  cells needing a change: {len(changed)}")
    for r, old, new in changed:
        print(f"    row {r}: {old} -> '{new}' (text)")

    if check_only:
        print("  (--check: no file written)")
        return 0
    if not changed:
        print("  Already canonical. Nothing to write.")
        return 0

    backup = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, backup)
    for r, _old, new in plan:
        cell = ws.cell(row=r, column=col_idx)
        cell.value = new
        cell.number_format = "@"  # force text, so a trailing zero can never be dropped again
    wb.save(path)
    print(f"  Backup written: {backup}")
    print(f"  Repaired and saved: {path}")
    return 0


def _same(ws, r: int, col_idx: int, new_id: str) -> bool:
    cell = ws.cell(row=r, column=col_idx)
    return str(cell.value).strip() == new_id and cell.data_type == "s"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--file", default="data/ISO.xlsx", help="ISO catalog workbook (relative to project root).")
    parser.add_argument("--column", default="ID", help="Name of the ID column header.")
    parser.add_argument("--check", action="store_true", help="Report what would change without writing.")
    args = parser.parse_args()

    path = (ROOT / args.file).resolve()
    if not path.exists():
        print(f"ERROR: file not found: {path}")
        sys.exit(2)
    sys.exit(repair_workbook(path, args.column, check_only=args.check))


if __name__ == "__main__":
    main()
